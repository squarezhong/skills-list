#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import math
import os
import re
import tempfile
from pathlib import Path
from urllib.parse import urlparse

os.environ.setdefault("MPLCONFIGDIR", str(Path(tempfile.gettempdir()) / "mplconfig-retail-index"))
os.environ.setdefault("XDG_CACHE_HOME", str(Path(tempfile.gettempdir()) / "cache-retail-index"))

import matplotlib

matplotlib.use("Agg")
import matplotlib.font_manager as fm
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


GOODS_CATEGORIES = [
    "粮油、食品类",
    "饮料类",
    "烟酒类",
    "服装、鞋帽、针纺织品类",
    "化妆品类",
    "金银珠宝类",
    "日用品类",
    "体育、娱乐用品类",
    "家用电器和音像器材类",
    "中西药品类",
    "文化办公用品类",
    "家具类",
    "通讯器材类",
    "石油及制品类",
    "汽车类",
    "建筑及装潢材料类",
]
RESTAURANT_CATEGORY = "限额以上单位餐饮收入"
ALL_CATEGORIES = [RESTAURANT_CATEGORY] + GOODS_CATEGORIES
EXCLUDE_TOKENS = {"", "不纳入", "剔除", "exclude", "excluded", "none", "nan"}


def normalize_label(value: object) -> str:
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return ""
    text = str(value).strip()
    text = text.replace("\u3000", "").replace(" ", "")
    text = text.replace("其中：", "")
    text = text.replace("服装鞋帽、针纺织品类", "服装、鞋帽、针纺织品类")
    text = text.replace("家用电器和音像器材", "家用电器和音像器材类")
    return text


def parse_number(value: object) -> float | None:
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "").replace("%", "")
    if text in {"", "-", "--", "—", "-\u3000", "nan"}:
        return None
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    return float(match.group(0)) if match else None


def flatten_columns(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out.columns = [str(col) if not isinstance(col, tuple) else " ".join(map(str, col)) for col in out.columns]
    return out


def read_source_tables(source: str) -> list[pd.DataFrame]:
    parsed = urlparse(source)
    suffix = Path(parsed.path if parsed.scheme else source).suffix.lower()
    if suffix in {".xls", ".xlsx"}:
        excel = pd.ExcelFile(source)
        return [pd.read_excel(source, sheet_name=sheet, header=None) for sheet in excel.sheet_names]
    if suffix in {".html", ".htm"} or parsed.scheme in {"http", "https"}:
        return [flatten_columns(table) for table in pd.read_html(source)]
    raise ValueError(f"Unsupported source type: {source}")


def extract_year_data(source: str, year: int) -> dict[str, tuple[float, float]]:
    found: dict[str, tuple[float, float]] = {}
    for raw_table in read_source_tables(source):
        table = raw_table.reset_index(drop=True)
        if table.empty:
            continue
        for _, row in table.iterrows():
            label = normalize_label(row.iloc[0] if len(row) else "")
            if label not in ALL_CATEGORIES:
                continue
            nums = [parse_number(value) for value in row.iloc[1:].tolist()]
            nums = [value for value in nums if value is not None]
            if len(row) >= 5 and parse_number(row.iloc[3]) is not None and parse_number(row.iloc[4]) is not None:
                amount = parse_number(row.iloc[3])
                yoy_pct = parse_number(row.iloc[4])
            elif len(nums) >= 2:
                amount = nums[-2]
                yoy_pct = nums[-1]
            else:
                raise ValueError(f"Could not identify amount/yoy for {year} {label} in {source}")
            if amount is None or yoy_pct is None:
                raise ValueError(f"Missing amount/yoy for {year} {label} in {source}")
            found[label] = (amount, yoy_pct / 100.0)
    if not found:
        raise ValueError(f"No NBS retail categories found in {source}")
    return found


def load_sources(path: Path) -> dict[int, str]:
    data = json.loads(path.read_text(encoding="utf-8"))
    return {int(year): str(source) for year, source in data.items()}


def truthy(value: object) -> bool:
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return False
    text = str(value).strip().lower()
    return text in {"1", "true", "yes", "y", "纳入", "include", "included"}


def load_mapping(path: Path) -> list[dict[str, object]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        df = pd.read_csv(path, dtype=str).fillna("")
    elif suffix in {".xls", ".xlsx"}:
        df = pd.read_excel(path, dtype=str).fillna("")
    elif suffix == ".json":
        data = json.loads(path.read_text(encoding="utf-8"))
        if isinstance(data, dict):
            data = data.get("constituents", [])
        df = pd.DataFrame(data).fillna("")
    else:
        raise ValueError(f"Unsupported mapping file: {path}")

    rename = {col: col.strip().lower() for col in df.columns}
    df = df.rename(columns=rename)
    aliases = {
        "股票代码": "code",
        "证券代码": "code",
        "公司名称": "name",
        "名称": "name",
        "指数权重": "weight",
        "权重": "weight",
        "修正分类": "category",
        "分类": "category",
        "是否纳入": "include",
        "纳入": "include",
        "备注": "note",
    }
    df = df.rename(columns={col: aliases.get(col, col) for col in df.columns})
    required = {"name", "category"}
    missing = required - set(df.columns)
    if missing:
        raise ValueError(f"Mapping file missing columns: {sorted(missing)}")

    records: list[dict[str, object]] = []
    for _, row in df.iterrows():
        category = str(row.get("category", "")).strip()
        include_value = row.get("include", "")
        include = truthy(include_value)
        if not str(include_value).strip() and category and category.lower() not in EXCLUDE_TOKENS:
            include = True
        if category.lower() in EXCLUDE_TOKENS:
            include = False
            category = "不纳入"
        if include and category not in ALL_CATEGORIES:
            raise ValueError(f"Unsupported category in mapping: {row.get('name')} -> {category}")
        records.append(
            {
                "code": str(row.get("code", "")).strip(),
                "name": str(row.get("name", "")).strip(),
                "weight": parse_number(row.get("weight", "")),
                "category": category,
                "include": include,
                "note": str(row.get("note", "")).strip(),
            }
        )
    return records


def build_rows(
    sources: dict[int, str],
    mapping: list[dict[str, object]],
    start_year: int,
    end_year: int,
) -> tuple[list[str], list[dict[str, object]], list[dict[str, object]]]:
    kept_categories = sorted(
        {str(row["category"]) for row in mapping if row["include"] and row["category"] in ALL_CATEGORIES},
        key=ALL_CATEGORIES.index,
    )
    if not kept_categories:
        raise ValueError("No included categories in mapping")

    category_notes = {category: [] for category in kept_categories}
    for row in mapping:
        if row["include"] and row["category"] in category_notes:
            category_notes[str(row["category"])].append(str(row["name"]))

    retail_rows: list[dict[str, object]] = []
    summary_rows: list[dict[str, object]] = []
    for year in range(start_year, end_year + 1):
        if year not in sources:
            raise ValueError(f"No source provided for {year}")
        year_data = extract_year_data(sources[year], year)
        included_for_year = []
        for category in kept_categories:
            amount_yoy = year_data.get(category)
            if amount_yoy is None:
                retail_rows.append(
                    {
                        "year": year,
                        "status": "纳入",
                        "category": category,
                        "amount": None,
                        "yoy": None,
                        "prior_base": None,
                        "source": sources[year],
                        "companies": "、".join(category_notes[category]),
                        "note": "该年官方发布表未列示该分类",
                    }
                )
                continue
            amount, yoy = amount_yoy
            prior_base = amount / (1.0 + yoy)
            item = {
                "year": year,
                "status": "纳入",
                "category": category,
                "amount": amount,
                "yoy": yoy,
                "prior_base": prior_base,
                "source": sources[year],
                "companies": "、".join(category_notes[category]),
                "note": "",
            }
            retail_rows.append(item)
            included_for_year.append(item)

        amount_sum = sum(float(row["amount"]) for row in included_for_year)
        base_sum = sum(float(row["prior_base"]) for row in included_for_year)
        missing = [
            row["category"]
            for row in retail_rows
            if row["year"] == year and row["status"] == "纳入" and row["amount"] is None
        ]
        summary_rows.append(
            {
                "year": year,
                "amount": amount_sum,
                "prior_base": base_sum,
                "yoy": amount_sum / base_sum - 1.0,
                "included_categories": len(included_for_year),
                "missing_categories": "、".join(missing),
            }
        )
    return kept_categories, retail_rows, summary_rows


def configure_font() -> None:
    candidates = ["PingFang SC", "Heiti SC", "Songti SC", "Arial Unicode MS", "Noto Sans CJK SC"]
    installed = {font.name for font in fm.fontManager.ttflist}
    for name in candidates:
        if name in installed:
            plt.rcParams["font.sans-serif"] = [name]
            break
    plt.rcParams["axes.unicode_minus"] = False


def make_chart(path: Path, title: str, years: list[int], amounts: list[float | None], yoys: list[float | None]) -> None:
    x = np.array(years)
    y = np.array([np.nan if value is None else float(value) for value in amounts], dtype=float)
    fig, ax = plt.subplots(figsize=(6.0, 3.2), dpi=160)
    ax.plot(x, y, color="#1F6F8B", linewidth=2.2, marker="o", markersize=5)
    ax.set_title(title, fontsize=11, pad=10)
    ax.set_xticks(x)
    ax.set_xlabel("年份")
    ax.set_ylabel("零售额（亿元）")
    ax.grid(True, axis="y", color="#D9E2EC", linewidth=0.8)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    finite = y[np.isfinite(y)]
    if len(finite):
        pad = (finite.max() - finite.min()) * 0.18 if finite.max() != finite.min() else max(finite.max() * 0.12, 100)
        ax.set_ylim(finite.min() - pad * 0.35, finite.max() + pad * 1.15)
    for year, amount, yoy in zip(years, amounts, yoys):
        if amount is None or yoy is None:
            continue
        ax.annotate(f"{float(yoy) * 100:.1f}%", (year, amount), textcoords="offset points", xytext=(0, 8), ha="center", fontsize=8, color="#17324D")
    fig.tight_layout()
    path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(path, bbox_inches="tight", facecolor="white")
    plt.close(fig)


def style_title(ws, last_col: int, title: str, subtitle: str) -> None:
    ws.sheet_view.showGridLines = False
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    ws.cell(1, 1).value = title
    ws.cell(1, 1).font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
    ws.cell(1, 1).fill = PatternFill("solid", fgColor="17324D")
    ws.cell(1, 1).alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
    ws.cell(2, 1).value = subtitle
    ws.cell(2, 1).font = Font(name="Arial", size=10, color="17324D")
    ws.cell(2, 1).fill = PatternFill("solid", fgColor="EAF2F8")
    ws.cell(2, 1).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 34


def add_table(ws, ref: str, name: str, percent_cols: set[int], amount_cols: set[int]) -> None:
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
    ws.add_table(table)
    header_row = int(re.findall(r"\d+", ref.split(":")[0])[0])
    for cell in ws[header_row]:
        if cell.value is not None:
            cell.font = Font(name="Arial", bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2E7D6F")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    max_row = int(re.findall(r"\d+", ref.split(":")[1])[0])
    for row in ws.iter_rows(min_row=header_row + 1, max_row=max_row):
        for cell in row:
            if cell.column in percent_cols:
                cell.number_format = "0.00%"
            elif cell.column in amount_cols:
                cell.number_format = "#,##0.0"


def autosize(ws, max_width: int = 48) -> None:
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        width = 10
        for row in range(1, ws.max_row + 1):
            value = ws.cell(row, col).value
            if value is not None:
                width = max(width, min(len(str(value)) + 2, max_width))
        ws.column_dimensions[letter].width = width


def write_workbook(
    output: Path,
    index_name: str,
    month: int,
    years: list[int],
    mapping: list[dict[str, object]],
    categories: list[str],
    retail_rows: list[dict[str, object]],
    summary_rows: list[dict[str, object]],
) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("成分股分类")
    style_title(ws, 6, f"{index_name} 成分股分类", "按消费端主营业务映射到国家统计局社零分类。")
    headers = ["代码", "公司名称", "权重", "修正分类", "是否纳入", "备注"]
    ws.append([])
    ws.append(headers)
    for row in mapping:
        ws.append([row["code"], row["name"], row["weight"], row["category"], "纳入" if row["include"] else "剔除", row["note"]])
    add_table(ws, f"A4:F{ws.max_row}", "ConstituentClassification", {3}, set())
    autosize(ws)

    ws = wb.create_sheet("原始限额以上数据")
    style_title(ws, 8, "限额以上社零原始数据", "来源为本地 NBS 文件或官方 stats.gov.cn 页面；金额单位为亿元。")
    headers = ["年份", "保留状态", "分类", f"1-{month}月绝对量", "原始同比", "反推上年基数", "来源", "对应指数内公司/说明"]
    ws.append([])
    ws.append(headers)
    for row in retail_rows:
        ws.append([row["year"], row["status"], row["category"], row["amount"], row["yoy"], row["prior_base"], row["source"], row["companies"] or row["note"]])
    add_table(ws, f"A4:H{ws.max_row}", "OriginalRetailData", {5}, {4, 6})
    autosize(ws)

    ws = wb.create_sheet("修正后汇总")
    last_col = 1 + len(years) * 2
    style_title(ws, last_col, f"修正后社会消费品零售总额（1-{month}月）", "分类纵向、年份横向；缺失的官方分类留空且不计入当年修正合计。")
    ws.freeze_panes = "B5"
    ws.cell(4, 1).value = "分类"
    for idx, year in enumerate(years):
        ws.cell(4, 2 + idx * 2).value = f"{str(year)[-2:]}年1-{month}月"
        ws.cell(4, 3 + idx * 2).value = "同比增长"
    lookup = {(row["year"], row["category"]): row for row in retail_rows}
    summary_lookup = {row["year"]: row for row in summary_rows}
    row_labels = categories + ["修正后社零总额"]
    for r, category in enumerate(row_labels, start=5):
        ws.cell(r, 1).value = category
        ws.cell(r, 1).font = Font(name="Arial", bold=(category == "修正后社零总额"))
        for idx, year in enumerate(years):
            if category == "修正后社零总额":
                amount, yoy = summary_lookup[year]["amount"], summary_lookup[year]["yoy"]
            else:
                amount, yoy = lookup[(year, category)]["amount"], lookup[(year, category)]["yoy"]
            ws.cell(r, 2 + idx * 2).value = amount
            ws.cell(r, 3 + idx * 2).value = yoy
            ws.cell(r, 2 + idx * 2).number_format = "#,##0.0"
            ws.cell(r, 3 + idx * 2).number_format = "0.00%"
    thin = Side(style="thin", color="D9E2EC")
    for row in ws.iter_rows(min_row=4, max_row=4 + len(row_labels), min_col=1, max_col=last_col):
        for cell in row:
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if cell.row == 4:
                cell.font = Font(name="Arial", bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="2E7D6F")
            elif cell.row == 4 + len(row_labels):
                cell.font = Font(name="Arial", bold=True)
                cell.fill = PatternFill("solid", fgColor="EAF2F8")
    ws.column_dimensions["A"].width = 30
    for col in range(2, last_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 13

    configure_font()
    chart_dir = output.with_suffix("")
    chart_dir.mkdir(parents=True, exist_ok=True)
    ws = wb.create_sheet("折线图")
    style_title(ws, 12, "各分类及修正后社零总额趋势", "横轴为年份，纵轴为零售额（亿元）；数据点标注同比增长。")
    anchors = ["A4", "G4", "A22", "G22", "A40", "G40", "A58", "G58", "A76", "G76", "A94", "G94", "A112", "G112"]
    chart_specs = []
    for category in categories:
        chart_specs.append((category, [lookup[(year, category)]["amount"] for year in years], [lookup[(year, category)]["yoy"] for year in years]))
    chart_specs.append(("修正后社零总额", [summary_lookup[year]["amount"] for year in years], [summary_lookup[year]["yoy"] for year in years]))
    for idx, (title, amounts, yoys) in enumerate(chart_specs):
        safe = re.sub(r"[\\/:*?\"<>|、，]+", "_", title)
        image_path = chart_dir / f"{idx + 1:02d}_{safe}.png"
        make_chart(image_path, title, years, amounts, yoys)
        img = XLImage(str(image_path))
        img.width = 560
        img.height = 310
        ws.add_image(img, anchors[idx])

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


def run_self_test(output: Path) -> None:
    with tempfile.TemporaryDirectory() as tmp:
        tmpdir = Path(tmp)
        sources: dict[int, str] = {}
        categories = [RESTAURANT_CATEGORY, "粮油、食品类", "饮料类", "体育、娱乐用品类"]
        for year in [2024, 2025]:
            rows = [["指标", "5月", None, f"1-{5}月", None], [None, "绝对量", "同比增长", "绝对量", "同比增长"]]
            for category in categories:
                amount = 1000 + (year - 2024) * 100 + categories.index(category) * 50
                yoy = 5 + categories.index(category)
                rows.append([category, None, None, amount, yoy])
            xls = tmpdir / f"{year}.xlsx"
            pd.DataFrame(rows).to_excel(xls, header=False, index=False)
            sources[year] = str(xls)
        mapping_path = tmpdir / "mapping.csv"
        pd.DataFrame(
            [
                {"code": "A", "name": "FoodCo", "weight": 0.5, "category": "粮油、食品类", "include": "纳入"},
                {"code": "B", "name": "DrinkCo", "weight": 0.5, "category": "饮料类", "include": "纳入"},
            ]
        ).to_csv(mapping_path, index=False)
        mapping = load_mapping(mapping_path)
        kept, retail, summary = build_rows(sources, mapping, 2024, 2025)
        write_workbook(output, "Self Test Index", 5, [2024, 2025], mapping, kept, retail, summary)
        print(json.dumps({"output": str(output), "kept_categories": kept, "summary": summary}, ensure_ascii=False, indent=2))


def main() -> None:
    parser = argparse.ArgumentParser(description="Build an index-adjusted NBS retail sales workbook.")
    parser.add_argument("--sources-json", type=Path, help="JSON mapping years to NBS source file paths or URLs.")
    parser.add_argument("--mapping", type=Path, help="CSV/XLSX/JSON constituent mapping with code,name,weight,category,include,note.")
    parser.add_argument("--month", type=int, default=5, help="YTD month number, e.g. 5 for 1-5月 or 6 for 1-6月.")
    parser.add_argument("--start-year", type=int)
    parser.add_argument("--end-year", type=int)
    parser.add_argument("--index-name", default="消费指数")
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--self-test", action="store_true", help="Run a synthetic smoke test and write --output.")
    args = parser.parse_args()

    if args.self_test:
        run_self_test(args.output)
        return
    if not args.sources_json or not args.mapping or not args.start_year or not args.end_year:
        parser.error("--sources-json, --mapping, --start-year, and --end-year are required unless --self-test is used")

    sources = load_sources(args.sources_json)
    mapping = load_mapping(args.mapping)
    years = list(range(args.start_year, args.end_year + 1))
    kept, retail, summary = build_rows(sources, mapping, args.start_year, args.end_year)
    write_workbook(args.output, args.index_name, args.month, years, mapping, kept, retail, summary)
    print(json.dumps({"output": str(args.output), "kept_categories": kept, "summary": summary}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
